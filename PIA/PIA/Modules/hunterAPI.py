from pyhunter import PyHunter
from openpyxl import Workbook
import shutil
import os


def Busqueda(hunter, organizacion):
    # Cantidad de resultados esperados de la búsqueda
    # El límite MENSUAL de Hunter es 50, cuidado!
    busqueda = 1
    resultado = hunter.domain_search(organizacion)
    return resultado


def GuardarInformacion(datosEncontrados, organizacion):
    # Se almacena la información obtenida desde la API
    # en una hoja de cálculo de Excel
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter-" + organizacion + ".xlsx")
    hoja.cell(1, 1, "Dominio")
    hoja.cell(1, 2, datosEncontrados['domain'])
    hoja.cell(2, 1, "Organización")
    hoja.cell(2, 2, datosEncontrados['organization'])
    hoja.cell(3, 1, "Correo")
    #hoja.cell(3, 2, datosEncontrados['emails'][0]['value'])
    hoja.cell(4, 1, "Nombre(s)")
    #hoja.cell(4, 2, datosEncontrados['emails'][0]['first_name'])
    hoja.cell(5, 1, "Apellidos")
    #hoja.cell(5, 2, datosEncontrados['emails'][0]['last_name'])
    libro.save("Hunter-" + organizacion + ".xlsx")

def hunter(link):
    hunter = PyHunter("9d9d6863cbebff140ed78630895b83e0768ee10d")
    datosEncontrados = Busqueda(hunter, link)
    if datosEncontrados is None:
        print("No hunter data")
        exit()
    else:
        #print(datosEncontrados)
        #print(type(datosEncontrados))
        GuardarInformacion(datosEncontrados, link)
    return hunter.account_information()

#hunter = PyHunter("9d9d6863cbebff140ed78630895b83e0768ee10d")
